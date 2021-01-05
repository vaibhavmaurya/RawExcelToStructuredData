'''

References:
    1. Read unstructured table from excel
    2. https://pbpython.com/pandas-excel-range.html
    3. https://openpyxl.readthedocs.io/en/stable/pandas.html

use following code to read from s3
Reference https://stackoverflow.com/questions/31416842/openpyxl-does-not-close-excel-workbook-in-read-only-mode

s3 = boto3.client('s3', aws_access_key_id=aws_id, aws_secret_access_key=aws_secret)
obj = s3.get_object(Bucket=bucket_name, Key=object_key)
data = obj['Body'].read()
df = pd.read_excel(io.BytesIO(data), encoding='utf-8')


import openpyxl
import io

xlsx_filename=r'C:/location/of/file.xlsx'
with open(xlsx_filename, "rb") as f:
    in_mem_file = io.BytesIO(f.read())

wb = openpyxl.load_workbook(in_mem_file, read_only=True)

'''
from collections import MutableMapping, Sequence
from functools import singledispatch
import pandas as pd
from skimage.measure import label, regionprops
import re
import openpyxl
from io import BytesIO
import os
import numpy as np

'''

'''


class CriteriaFieldNotFound(Exception):
    pass


class InvalidParameter(Exception):
    pass


class ExcelFileNotFound(Exception):
    pass


class WorkbookLoadFailed(Exception):
    pass


class ExcelSheetNotFound(Exception):
    pass


class ExcelSheetNotFoundForPattern(Exception):
    pass


@singledispatch
def resolve_index(key):
    # print(key)
    return key, key


@resolve_index.register(int)
def _(key):
    return key, key


@resolve_index.register(slice)
def _(key):
    return key.start, key.stop


@resolve_index.register(Sequence)
def _(key):
    if not validate_index(key):
        raise IndexError(f'Not a valid {key} indices sequence')
    return min(key), max(key)


@singledispatch
def resolve_index_tuple(key):
    return 1


@resolve_index_tuple.register(int)
def _(key):
    return 0


@resolve_index_tuple.register(slice)
def _(key):
    return slice(0, key.stop - key.start if key.stop else None, key.step)


@resolve_index_tuple.register(Sequence)
def _(key):
    n = min(key)
    return [x - n for x in key]


def validate_index(seq: Sequence, obj_type: object = int, is_any: bool = False) -> bool:
    if is_any:
        any_or_all = any
    else:
        any_or_all = all
    if not isinstance(seq, Sequence):
        return False
    return any_or_all((isinstance(x, obj_type) for x in seq))


class ExcelSheet:
    def __init__(self, sheet):
        self.sheet = sheet

    def __resolve_xory(self, key: slice, is_row=True):
        start, stop = key.start, key.stop
        if not start:
            start = 1
        if not stop:
            if not is_row:
                stop = self.sheet.max_column
            else:
                stop = self.sheet.max_row
        return start, stop

    def __resolve_row_column(self, indices, is_row=True):
        a, max_val = ('row', self.sheet.max_row) if is_row else ("columns", self.sheet.max_column)
        min_len, max_len = resolve_index(indices)
        if not min_len:
            min_len = 1
        if not max_len:
            max_len = max_val
        if min_len < 1 or max_len > max_val:
            raise IndexError(
                f'Indices out of boundary for {a} where indices are min_len:{min_len} max_len:{max_len} max_val:{max_val} indices:{indices}')
        return min_len, max_len

    def __getitem__(self, item):
        a = []
        is_seq_access = isinstance(item, tuple) and \
                        ((len(item) > 2 and validate_index(item)) or
                         (len(item) == 2 and validate_index(item, obj_type=Sequence, is_any=True)) or
                         (len(item) == 1 and validate_index(item[0]))
                         )
        index_i, index_j = None, None
        if isinstance(item, tuple) and len(item) > 2:
            if not validate_index(item):
                raise IndexError(f"Not a valid index: {item}")
            min_row, max_row = self.__resolve_row_column(item)
            min_col, max_col = 1, self.sheet.max_column
        else:
            try:
                index_i, index_j = item
            except (ValueError, TypeError):
                index_i = item[0] if isinstance(item, tuple) else item

            min_row, max_row = self.__resolve_row_column(index_i) if index_i else (1, self.sheet.max_row)
            ## is_row=False was missing for columns index
            min_col, max_col = self.__resolve_row_column(index_j, is_row=False) if index_j else (
            1, self.sheet.max_column)

            if (min_row == max_row and min_col == max_col) and isinstance(min_row, int) and isinstance(min_col, int):
                return str(self.sheet.cell(min_row, min_col).value).strip()

        b = self.sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)

        for x in b:
            c = []
            for y in x:
                c.append(str(y.value).strip())
            a.append(c)
        g = np.array(a, dtype='str')

        if is_seq_access:
            if len(item) > 2:
                g = g[[k - min_row for k in item],]
            elif len(item) == 1:
                g = g[[k - min_row for k in item[0]],]
            else:
                w, z = resolve_index_tuple(item[0]), resolve_index_tuple(item[1])
                # print(w,z)
                if w is None:
                    g = g[:, z]
                elif z is None:
                    g = g[w]
                elif w == 0 and z == 0:
                    g = g[0, 0]
                else:
                    try:
                        g = g[w][:, z]
                    except:
                        g = g[w, z]
                # g = g[:, z] if w is None else g[w] if z is None else g[0, 0] if w == 0 and z == 0 else g[x][:, y]
        return g

    def scan_sheet(self, set_of_patterns: set, min_row=1, max_row=None, min_col=1, max_col=None, partial_match=True,
                   match_threshold=0):
        match_completed = False

        if not max_row:
            max_row = self.sheet.max_row
        if not max_col:
            max_col = self.sheet.max_column
        if not set_of_patterns or not isinstance(set_of_patterns, set):
            raise Exception('pattern param of set type is mandatory')
        if min_row > max_row or min_col > max_col:
            raise Exception('row and col range values are not correct')
        matched_items = {x: None for x in set_of_patterns}
        b = self.sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col,
                                 max_col=max_col)

        # re.sub('[^A-Za-z0-9]+', '', mystring)
        check = 0
        for x in b:
            if match_completed:
                break
            for y in x:
                if not y.value:
                    continue
                if len(set_of_patterns) == 0:
                    match_completed = True
                    break
                for item in set_of_patterns:
                    l = re.sub("[()?$\\\/,\*\+]", '', str(item).lower())
                    p = re.sub("[()?$\\\/,\*\+]", '', str(y.value).lower())
                    g = re.search(re.compile(l), p) if partial_match else re.search(
                        re.compile(str(item)), str(y.value))
                    if g:
                        matched_items[item] = {"coordinate": (y.row, y.col_idx), "cell_value": y.value}
                        set_of_patterns.remove(item)
                        break
        return matched_items


class ExcelWorkbook:

    def __init__(self, filename: str=None, bytesIO=None):
        if filename is None and bytesIO is None:
            raise InvalidParameter(filename)
        if bytesIO is None and not os.path.exists(filename):
            raise ExcelFileNotFound(filename)
        try:
            if bytesIO is not None:
                self.excelBytes = bytesIO
                self.__wb = openpyxl.load_workbook(self.excelBytes, data_only=True)
            else:
                with open(filename, "rb") as f:
                    self.excelBytes = BytesIO(f.read())
                    self.__wb = openpyxl.load_workbook(self.excelBytes, data_only=True)
        except Exception as e:
            raise WorkbookLoadFailed(e)

    def get_sheet_names(self):
        return self.__wb.sheetnames

    def close_workbook(self):
        self.__wb.close()

    def get_sheet(self, sheet_name=None, pattern=None):
        if not (sheet_name or pattern) or not (isinstance(sheet_name, str) or isinstance(pattern, list)):
            raise Exception('Any one param is manadatory with string type or list type')

        # print(f" Inside  get_sheet {sheet_name} for {pattern} inside {self.__wb.sheetnames}")
        if sheet_name:
            try:
                return ExcelSheet(self.__wb[sheet_name]), sheet_name
            except Exception as e:
                raise ExcelSheetNotFound(e)

        if len(pattern) > 0:
            k = re.compile('|'.join(pattern))
            # print(self.__wb.sheetnames)
            for sheetname in self.__wb.sheetnames:
                if k.search(sheetname.lower()):
                    # print(f"Found sheet for {sheetname} for {pattern}")
                    return ExcelSheet(self.__wb[sheetname]), sheetname
            raise ExcelSheetNotFoundForPattern(str(pattern))


class BuildDataFrame:

    def __init__(self, filename: str, bytesIO=None):
        self._wb = ExcelWorkbook(filename, bytesIO)
        self.filename = filename
        self.ID = re.sub(r"[^a-zA-Z0-9]", "", filename[filename.rindex('\\') + 1: filename.rindex('.xlsx')])

    def __repr__(self):
        return f"DataFrame({self._sheet})"

    def __str__(self):
        return f"DataFrame({self._sheet})"

    def set_sheet(self, sheet_name=None, pattern=None):
        if sheet_name is None and pattern is None:
            # print(f"sheet_name: {sheet_name} pattern:{pattern}")
            sheet_name = self._wb.get_sheet_names()[0]
        self._sheet, self.sheetname = self._wb.get_sheet(sheet_name, pattern)
        # print(f" YOUR Sheet is {self._sheet.sheet} for the pattern={pattern}")
        self.regiongroups = []
        df = pd.DataFrame(list(self._sheet.sheet.values))
        # this basically converts your table into 0s and 1s where 0 is NaN and 1 for non NaN
        self.labels = label(np.array(df.notnull().astype('int')))
        self.skip_regions = []
        for s in regionprops(self.labels):
            self.regiongroups.append(s)

    def extract_excel_table(self, i_header, found_columns, column_mappings=None):
        column_mappings = column_mappings or {}
        if not i_header and not found_columns:
            raise InvalidParameter(" i_header, j_header, found_columns ")

        # sheetname
        # found_cols = [(column_mappings.get(key, key), *val["coordinate"]) for key, val in found_columns.items()]
        # found_cols.sort(key=lambda x: x[2])
        # print((i_header + 1), [j[2] for j in found_cols])
        # data = self._sheet[(i_header + 1):, [j[2] for j in found_cols]]
        # print(data.shape)
        # df = pd.DataFrame(data, columns=[x[0] for x in found_cols])
        print("Extract Direct Table")
        df = None

        #  use feature usecols so that limited columns will be extracted here
        # df = pd.read_excel(self.filename, sheet_name=sheetname, skiprows=i_header-1)
        print(f"Header is : {i_header} and sheet name is {self.sheetname}")
        # print(column_mappings)
        # print(found_columns)
        df = pd.read_excel(self._wb.excelBytes, usecols=list(column_mappings.keys()),
                           sheet_name=self.sheetname,
                           skiprows=i_header - 1, encoding='utf-8')

        if column_mappings is not None:
            df.rename(columns=column_mappings, inplace=True)
            df = df[[v for _, v in column_mappings.items()]]

        df = df.assign(ID=[self.ID] * df.shape[0])
        df = df.assign(row_number=[i for i in range(i_header + 1, df.shape[0] + i_header + 1)])
        return df

    # TODO: Do it later
    def check_stop_criteria(self, end_row: int, end_col: int, start_row: int = 1, start_col: int = 1, criteria=None):
        if not end_row:
            end_row = self._sheet.sheet.max_row
        if not end_col:
            end_col = self._sheet.sheet.max_column

    def _find_region(self, lookup_pattern, match_threshold=0):
        anchor = None
        # print(self.skip_regions)
        # regions = ((i, val) for i,val in enumerate(self.regiongroups) if i not in self.skip_regions)
        for i, region in enumerate(self.regiongroups):
            min_row, min_col, max_row, max_col = (x + 1 for x in region.bbox)
            anchor = self._sheet.scan_sheet({lookup_pattern}, min_row=min_row, max_row=max_row, min_col=min_col,
                                            max_col=max_col, partial_match=True, match_threshold=match_threshold)
            if anchor.get(lookup_pattern) is not None and anchor[lookup_pattern].get("coordinate", "@") != "@":
                return anchor, region, i
        raise Exception(f"Table not found for the lookup pattern {lookup_pattern}")

    ## Here lookup_pattern is a set of patterns and minimum match_threshold must match
    def _find_region_by_mode(self, lookup_pattern, match_threshold=1, row_vs_col="row"):
        anchor = None
        # print(self.skip_regions)
        # regions = ((i, val) for i,val in enumerate(self.regiongroups) if i not in self.skip_regions)
        sort_criteria = 0 if row_vs_col == "row" else 1
        if match_threshold > len(lookup_pattern):
            match_threshold = len(lookup_pattern)

        for i, region in enumerate(self.regiongroups):
            min_row, min_col, max_row, max_col = (x + 1 for x in region.bbox)
            anchor = self._sheet.scan_sheet(lookup_pattern, min_row=min_row, max_row=max_row, min_col=min_col,
                                            max_col=max_col, partial_match=True)
            found_items = {k:v for k, v in anchor.items() if v is not None}
            # print(anchor)
            if len(found_items) >= match_threshold:
                return found_items, region, i
        raise Exception(f"Table not found for the lookup pattern {lookup_pattern}")

    def find_table(self, lookup_pattern, column_mappings=None, value_mappings=None,
                   partial_match=True, excel_table=False):
        if not isinstance(lookup_pattern, str):
            raise InvalidParameter("lookup_pattern")
        if column_mappings and not isinstance(column_mappings, MutableMapping):
            raise InvalidParameter("column_mappings")
        if value_mappings and not isinstance(value_mappings, MutableMapping):
            raise InvalidParameter("column_mappings")
        # find table anchor

        anchor, _, _ = self._find_region(lookup_pattern)

        print(anchor)
        if anchor[lookup_pattern].get("coordinate", "@") == "@":
            raise Exception(f"Table not found for the lookup pattern {lookup_pattern}")

        i_header, j_header = anchor[lookup_pattern].get("coordinate")
        # print("hello there")
        # print(i_header, j_header)
        found_columns = self._sheet.scan_sheet(set(column_mappings.keys()), min_row=i_header,
                                               max_row=i_header, min_col=1,
                                               partial_match=partial_match)
        # print(found_columns)
        found_columns = {key: val for key, val in found_columns.items() if val is not None}
        if not any(x.get("coordinate", "@") != "@" for _, x in found_columns.items()):
            raise Exception("Table not found for the given column_mappings")

        # column_mappings = {key:column_mappings[key] for key, val in found_columns.items() if val is not None}
        # cell_value
        new_column_mapping = {val["cell_value"]: column_mappings[key] for key, val in found_columns.items() if
                              val is not None}

        if excel_table:
            print("Directly read dataframe")
            # print(new_column_mapping)

        df = self.extract_excel_table(i_header,
                                      found_columns,
                                      new_column_mapping)
        return df, list(new_column_mapping.values())

    ## TODO: Discard found regions
    ## Skip rows is provided for performance reason, to reduce scan area
    ## stop search if n-search_stop_at fields are found
    def search_regions_form_values(self, search_pattern: set, search_stop_at=3):
        # self._sheet self.regiongroups self.labels self.skip_regions
        final_fields = {}
        found_col, region, i = self._find_region_by_mode(search_pattern, match_threshold=search_stop_at, row_vs_col="row")

        self.skip_regions.append(i)
        # min_row, min_col, max_row, max_col = (x + 1 for x in region.bbox)
        # Need to to scan since all we got back already in achor
        # found_col = self._sheet.scan_sheet(search_pattern, min_row, max_row, min_col, max_col)

        ## The region width can be decieving for example what if max_col=1. Though one can rely on found rows
        ## So I am setting the max_col to 5 for safe side, we can ponder about it later
        ## This decision will not hurt the performance
        max_col_offset = 5
        # print(found_col)
        for k, val in found_col.items():
            if val is not None:
                a, b = val["coordinate"]
                # array starts at 0 though excel starts at 1
                # print(f"{k}: ({a},{b}) - maxcol:{max_col-1}")
                # print(self.labels.shape)
                # print(self.labels[a-1, b:max_col-1])
                # print(np.nonzero(self.labels[a-1, b:max_col-1]))
                max_col = b + max_col_offset if (b + max_col_offset) < self.labels.shape[1] else self.labels.shape[1]

                non_zero = np.nonzero(self.labels[a - 1, b:max_col - 1])[0]

                if len(non_zero) > 0:
                    val_col = b + 1 + non_zero[0]
                    # print(f"{k}: ({a},{b}) - val: {val_col}")
                    final_fields[k] = self._sheet[a, val_col][0][0]
                else:
                    final_fields[k] = ''
        return final_fields
